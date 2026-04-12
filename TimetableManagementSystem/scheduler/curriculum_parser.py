"""
Upload parser for Curriculum Excel file.
Multi-sheet .xlsx — one sheet per programme (sheet name = programme name or code).

Columns per sheet: S/N, Course_Code, Course_Name, Study_Period
  - S/N          : serial number — ignored
  - Course_Code  : unique course code
  - Course_Name  : course title/name
  - Study_Period : combined year+semester string, e.g. "Year 1:Semester 1"
                   Parsed to extract study_year and semester integers.

is_exam and is_lab default to False (update via other means if needed).
"""
import re
import openpyxl
from .models import Programme, Course

COLUMN_ALIASES = {
    'sn':           ['s/n', 'sn', 'no', 'number', '#', 's_n', 'serial'],
    'course_code':  ['course_code', 'code', 'course code', 'subject_code', 'module_code', 'coursecode'],
    'course_name':  ['course_name', 'name', 'course_title', 'course title', 'title', 'subject', 'module', 'course name'],
    'study_period': ['study_period', 'period', 'study_period', 'study period', 'year_semester',
                     'year/semester', 'year_and_semester', 'term', 'studyperiod',
                     'study_year_semester', 'year_sem', 'year_sem_', 'period_of_study'],
}


def _normalize(value):
    if not value and value != 0:
        return ''
    s = str(value).strip().lower()
    s = re.sub(r'[\s_]+', '_', s)   # collapse all whitespace variants to underscore
    s = re.sub(r'[^\w]', '', s)     # drop remaining non-word characters
    return s


def _detect_columns(headers):
    col_map = {}
    normalized = [_normalize(h) for h in headers]
    for canonical, aliases in COLUMN_ALIASES.items():
        norm_aliases = {_normalize(a) for a in aliases}
        for idx, h in enumerate(normalized):
            if h in norm_aliases:
                col_map[canonical] = idx
                break
    return col_map


def _parse_study_period(value):
    """
    Parse a study period string into (study_year, semester).
    Accepts formats like:
      "Year 1:Semester 1", "Year 1: Semester 1", "year1 sem2",
      "Y1S1", "1:1", "Year 2 Semester 1"
    Returns (study_year: int, semester: int).
    Raises ValueError if parsing fails.
    """
    s = str(value).lower().strip()

    # Try explicit "year X ... sem Y" pattern first
    year_match = re.search(r'year\s*(\d)', s)
    sem_match = re.search(r'sem(?:ester)?\s*(\d)', s)

    if year_match and sem_match:
        return int(year_match.group(1)), int(sem_match.group(1))

    # Try "Y1S1" shorthand
    short_match = re.match(r'y(\d)s(\d)', s.replace(' ', ''))
    if short_match:
        return int(short_match.group(1)), int(short_match.group(2))

    # Try plain "X:Y" or "X/Y"
    num_match = re.match(r'(\d)\s*[:\/]\s*(\d)', s)
    if num_match:
        return int(num_match.group(1)), int(num_match.group(2))

    raise ValueError(f"Cannot parse study period: '{value}'")


def parse_curriculum(file_obj):
    """
    Parse multi-sheet curriculum Excel and create/update Course records.
    Each sheet name must match an existing Programme name or code (case-insensitive).

    Returns:
        {'success_count': int, 'errors': [str], 'sheets_processed': [str]}
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        return {'success_count': 0, 'errors': [f"Cannot open file: {e}"], 'sheets_processed': []}

    # Build a lookup: normalized name/code -> Programme object
    all_programmes = list(Programme.objects.select_related('school').all())
    prog_lookup = {}
    for p in all_programmes:
        prog_lookup[p.name.lower().strip()] = p
        prog_lookup[p.code.lower().strip()] = p
        # Also index the bare name without the "(CODE)" suffix, e.g. "Architecture"
        bare = re.sub(r'\s*\(.*?\)\s*$', '', p.name).strip().lower()
        if bare:
            prog_lookup[bare] = p

    def _fuzzy(s):
        """Strip punctuation and spaces for loose comparison: 'B.A. CDS' -> 'bacds'"""
        return re.sub(r'[^a-z0-9]', '', s.lower())

    def _match_programme(sheet):
        # Strip regular and non-breaking whitespace
        key = sheet.replace('\xa0', ' ').strip().lower()

        # Strip trailing year suffix, e.g. "BSc. HIP 2019" -> "BSc. HIP"
        key_no_year = re.sub(r'\s*\b(19|20)\d{2}\b\s*$', '', key).strip()

        # 1. Exact match (with and without year suffix)
        for k in (key, key_no_year):
            if k in prog_lookup:
                return prog_lookup[k]

        # 2. Fuzzy match: strip all punctuation/spaces, compare
        fuzzy_key = _fuzzy(key_no_year)
        for p in all_programmes:
            if fuzzy_key == _fuzzy(p.code) or fuzzy_key == _fuzzy(p.name):
                return p

        # 3. Sheet name contains programme code as a word token
        for p in all_programmes:
            code_pat = re.escape(p.code.lower())
            if re.search(r'(?<![a-z])' + code_pat + r'(?![a-z])', key_no_year):
                return p

        # 4. Fuzzy partial: programme code is contained in the fuzzy sheet name
        for p in all_programmes:
            if _fuzzy(p.code) in fuzzy_key:
                return p

        # 5. Programme name/code contains the stripped sheet name (partial)
        for p in all_programmes:
            if key_no_year in p.name.lower() or key_no_year in p.code.lower():
                return p

        return None

    total_success = 0
    all_errors = []
    sheets_processed = []
    available_programmes = ', '.join(
        f"'{p.code}' / '{p.name}'" for p in all_programmes
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            all_errors.append(f"Sheet '{sheet_name}': empty — skipped")
            continue

        # Match sheet name to a Programme
        programme = _match_programme(sheet_name)
        if not programme:
            all_errors.append(
                f"Sheet '{sheet_name}': no Programme found matching this name or code — skipped. "
                f"Available: {available_programmes}"
            )
            continue

        # Find the actual header row — skip title/merged rows at the top.
        # Scan up to the first 10 rows; prefer a row that has ALL 3 required columns,
        # fall back to a row with at least course_code + course_name if nothing better found.
        header_row_idx = None
        col_map = {}
        partial_idx = None
        partial_map = {}
        for scan_idx, scan_row in enumerate(rows[:10]):
            candidate = _detect_columns(scan_row)
            has_code = 'course_code' in candidate
            has_name = 'course_name' in candidate
            has_period = 'study_period' in candidate
            if has_code and has_name and has_period:
                header_row_idx = scan_idx
                col_map = candidate
                break
            if has_code and has_name and partial_idx is None:
                partial_idx = scan_idx
                partial_map = candidate

        if header_row_idx is None:
            # Use best partial match if found, else row 0
            if partial_idx is not None:
                header_row_idx = partial_idx
                col_map = partial_map
            else:
                # Default to row 0 — the data-driven fallback below may still rescue it
                header_row_idx = 0
                col_map = _detect_columns(rows[0])

        # ── Data-driven fallback: if course_name is missing but course_code is found,
        # check whether the course_code column actually holds long descriptive names
        # (mismatched header). If so, treat the S/N column as course_code instead.
        # This handles sheets where columns are labelled in wrong order, e.g.:
        #   ['Programme', 'S/N', 'Course Code', 'Study Period']
        # where 'S/N' really holds codes and 'Course Code' really holds names.
        if 'course_name' not in col_map and 'course_code' in col_map:
            data_rows = rows[header_row_idx + 1: header_row_idx + 6]
            code_col_vals = [str(r[col_map['course_code']]) for r in data_rows if r[col_map['course_code']]]
            # If the "code" column values look like long names (avg length > 15), it's actually the name column
            if code_col_vals and sum(len(v) for v in code_col_vals) / len(code_col_vals) > 15:
                # Reassign: code_col → course_name; look for real code in sn or first unused column
                col_map['course_name'] = col_map.pop('course_code')
                # Find course_code: prefer 'sn' column (often misused for codes), else first unused col
                if 'sn' in col_map:
                    col_map['course_code'] = col_map.pop('sn')
                else:
                    used = set(col_map.values())
                    for ci in range(len(rows[header_row_idx])):
                        if ci not in used:
                            col_map['course_code'] = ci
                            break

        required = ['course_code', 'course_name', 'study_period']
        missing = [r for r in required if r not in col_map]
        if missing:
            # Show each of the first 5 rows so the user can see what headers are present
            sample = '; '.join(
                f"Row {i+1}: {[str(c) for c in rows[i] if c is not None]}"
                for i in range(min(5, len(rows)))
            )
            all_errors.append(
                f"Sheet '{sheet_name}': missing columns {missing}. "
                f"Header scan (first 5 rows): {sample}"
            )
            continue

        sheet_success = 0
        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            try:
                course_code = str(row[col_map['course_code']]).strip() if row[col_map['course_code']] else None
                course_name = str(row[col_map['course_name']]).strip() if row[col_map['course_name']] else None
                study_period_raw = row[col_map['study_period']]

                if not course_code or course_code.lower() in ('none', 'null', ''):
                    all_errors.append(f"Sheet '{sheet_name}' Row {row_idx}: missing course code — skipped")
                    continue
                if not course_name or course_name.lower() in ('none', 'null', ''):
                    all_errors.append(f"Sheet '{sheet_name}' Row {row_idx}: missing course name — skipped")
                    continue
                if study_period_raw is None or str(study_period_raw).strip() == '':
                    all_errors.append(f"Sheet '{sheet_name}' Row {row_idx}: missing study period — skipped")
                    continue

                try:
                    study_year, semester = _parse_study_period(study_period_raw)
                except ValueError as ve:
                    all_errors.append(f"Sheet '{sheet_name}' Row {row_idx}: {ve}")
                    continue

                if semester not in (1, 2):
                    all_errors.append(f"Sheet '{sheet_name}' Row {row_idx}: semester must be 1 or 2, got '{semester}'")
                    continue
                if not (1 <= study_year <= 6):
                    all_errors.append(f"Sheet '{sheet_name}' Row {row_idx}: study_year must be 1-6, got '{study_year}'")
                    continue

                Course.objects.update_or_create(
                    code=course_code,
                    programme=programme,
                    section='',
                    defaults={
                        'title': course_name,
                        'semester': semester,
                        'study_year': study_year,
                        'is_exam': False,
                        'is_lab': False,
                    }
                )
                sheet_success += 1

            except ValueError as ve:
                all_errors.append(f"Sheet '{sheet_name}' {ve}")
            except Exception as e:
                all_errors.append(f"Sheet '{sheet_name}' Row {row_idx}: {e}")

        total_success += sheet_success
        sheets_processed.append(f"{sheet_name} ({sheet_success} courses)")

    return {
        'success_count': total_success,
        'errors': all_errors,
        'sheets_processed': sheets_processed,
    }
