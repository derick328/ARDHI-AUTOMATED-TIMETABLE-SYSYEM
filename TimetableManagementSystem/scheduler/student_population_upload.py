"""
Upload parser for Student Populations Excel file.
Multi-sheet .xlsx — one sheet per school/faculty (sheet name = school name).
Columns: Programme_Code, Programme_Name, Study_Year, Male_Count, Female_Count, Total_Count

Notes:
  - Study_Year must be integer (1, 2, 3, 4), NOT Roman numerals
  - Rows with Total_Count = 0 are skipped
  - Creates/updates School, Programme, and StudentCount records
"""
import openpyxl
from .models import School, Programme, StudentCount

COLUMN_ALIASES = {
    'programme_code': ['programme_code', 'prog_code', 'code', 'programme code', 'program_code'],
    'programme_name': ['programme_name', 'programme', 'program_name', 'program', 'name', 'programme name'],
    'study_year':     ['study_year', 'year', 'year_of_study', 'level', 'stage', 'study year'],
    'male_count':     ['male_count', 'male', 'males', 'boys', 'male count'],
    'female_count':   ['female_count', 'female', 'females', 'girls', 'female count'],
    'total_count':    ['total_count', 'total', 'total students', 'students', 'count'],
}


def _normalize(value):
    return str(value).strip().lower().replace(' ', '_') if value else ''


def _detect_columns(headers):
    col_map = {}
    normalized = [_normalize(h) for h in headers]
    for canonical, aliases in COLUMN_ALIASES.items():
        for idx, h in enumerate(normalized):
            if h in aliases:
                col_map[canonical] = idx
                break
    return col_map


def _parse_int(value, default=0):
    try:
        v = str(value).strip()
        if not v or v.lower() in ('none', 'null', ''):
            return default
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _roman_to_int(value):
    """Convert Roman numeral string to integer, fallback to int parse."""
    roman_map = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5}
    v = str(value).strip().upper()
    if v in roman_map:
        return roman_map[v]
    return _parse_int(value)


def parse_student_populations(file_obj):
    """
    Parse multi-sheet student population Excel.

    Returns:
        {'success_count': int, 'errors': [str], 'sheets_processed': [str]}
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        return {'success_count': 0, 'errors': [f"Cannot open file: {e}"], 'sheets_processed': []}

    total_success = 0
    all_errors = []
    sheets_processed = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            all_errors.append(f"Sheet '{sheet_name}': empty — skipped")
            continue

        # Get or create School from sheet name
        school, _ = School.objects.get_or_create(
            name__iexact=sheet_name.strip(),
            defaults={'name': sheet_name.strip(), 'code': sheet_name.strip()[:20].upper()}
        )
        if not school.pk:
            school, _ = School.objects.get_or_create(name=sheet_name.strip())

        headers = rows[0]
        col_map = _detect_columns(headers)

        required = ['programme_code', 'study_year']
        missing = [r for r in required if r not in col_map]
        if missing:
            all_errors.append(
                f"Sheet '{sheet_name}': missing columns {missing}. Found: {list(headers)}"
            )
            continue

        sheet_success = 0
        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                prog_code_raw = row[col_map['programme_code']]
                prog_code = str(prog_code_raw).strip().upper() if prog_code_raw else None

                if not prog_code or prog_code.lower() in ('none', 'null', ''):
                    continue  # blank row — silently skip

                study_year_raw = row[col_map['study_year']]
                study_year = _roman_to_int(study_year_raw)

                if not (1 <= study_year <= 6):
                    all_errors.append(
                        f"Sheet '{sheet_name}' Row {row_idx}: invalid study_year '{study_year_raw}' — skipped"
                    )
                    continue

                male_count = _parse_int(row[col_map.get('male_count', -1)]) if 'male_count' in col_map else 0
                female_count = _parse_int(row[col_map.get('female_count', -1)]) if 'female_count' in col_map else 0
                total_count = _parse_int(row[col_map.get('total_count', -1)]) if 'total_count' in col_map else 0

                if total_count == 0:
                    total_count = male_count + female_count

                # Skip rows with zero students
                if total_count == 0:
                    continue

                prog_name_raw = row[col_map.get('programme_name', -1)] if 'programme_name' in col_map else None
                prog_name = str(prog_name_raw).strip() if prog_name_raw else prog_code

                programme, _ = Programme.objects.update_or_create(
                    code=prog_code,
                    defaults={'name': prog_name, 'school': school}
                )

                StudentCount.objects.update_or_create(
                    programme=programme,
                    study_year=study_year,
                    defaults={
                        'male_count': male_count,
                        'female_count': female_count,
                        'total_count': total_count,
                    }
                )
                sheet_success += 1

            except Exception as e:
                all_errors.append(f"Sheet '{sheet_name}' Row {row_idx}: {e}")

        total_success += sheet_success
        sheets_processed.append(f"{sheet_name} ({sheet_success} records)")

    return {
        'success_count': total_success,
        'errors': all_errors,
        'sheets_processed': sheets_processed,
    }
